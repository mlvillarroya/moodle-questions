#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl, codecs, sys, re
from tkinter import Tk,Label,Frame,Button,Text,END
from tkinter.ttk import Entry
from tkinter.filedialog import askopenfilename
from tkinter.messagebox import showerror
import webbrowser

# def load_file(cuadro):
#     fname = askopenfilename(filetypes=(("Excel file", "*.xlsx"),))
#     if fname:
#         try:
#             ruta_archivo.config(text=fname)
#             cuadro.insert(END, "Archivo cargado correctamente\n")
#         except:  # <- naked except is a bad idea
#             showerror("Open Source File", "Failed tofruski read file\n'%s'" % fname)
#         return

def generar(archivo, cuadro_texto,porcentaje):
    cuadro_texto.insert(END,"Procedemos a abrir el archivo\n")
    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
    except IOError:
        cuadro_texto.insert(END,u'\nNo existe el archivo Excel llamado \"' + archivo + u'\".\nPor favor, vuélvelo a intentar\n\n\n')     
        #cuadro_texto.insert(END,u'\nNo existe el archivo Excel llamado \"' + archivo + u'\". Saliendo del programa\nPara más información escribir generar --help')
        #sys.exit()
    else:
        cuadro_texto.insert(END,"Archivo abierto correctamente, procedemos a crear las preguntas\n")

    #Primera pestaña: contiene el nombre del ejercicio y las preguntas de opción
    try:
        sheet = wb['preguntas_opcion']
    except KeyError:
        cuadro_texto.insert(END,u'\nNo existe la pestaña \"preguntas_opcion\", no se añadirán preguntas de este tipo')
    else:
        # Primera hoja: preguntas de opción. hola caracola
        clase=sheet['B1'].value
        tema=sheet['D1'].value
        codigo=clase+'_'+tema
        cuadro_texto.insert(END,u'\nNombre del ejercicio: ' + codigo)

        # Creación del archivo de salida
        resultFile = codecs.open(codigo + ".txt", "w", "utf-8")
        cuadro_texto.insert(END,u'\n    Creado el archivo ' + codigo + ".txt")

        cuadro_texto.insert(END,u'\n\nLeyendo preguntas de opción')

        resultFile.write('$CATEGORY: $course$/' + tema + '\n\n')

        for row in range(3, sheet.max_row + 1):
            if porcentaje:
                signo_positivo = ' ~%100%'
                signo_negativo = ' ~%-'+porcentaje+'%'
            else:
                signo_positivo = ' ='
                signo_negativo = ' ~'
            # En cada fila, vamos leyendo las diferentes celdas
            celda=[]
            #memorizamos toda la fila
            for column in range(1, 8):
                celda.append((sheet.cell(row,column).value))
            if (celda[6]):
                resultFile.write('$CATEGORY: $course$/' + tema + '/' + celda[6] + '\n\n')
            if (celda[0]):
                #Pregunta. Campo obligatorio
                pregunta='::_opcion_' + str(row-2).zfill(2) + '_' + celda[0][0:9] + '::' + celda[0] + ' {\n'
                #Respuesta correcta. Campo obligatorio
                if (celda[1]):
                    pregunta=pregunta + signo_positivo + celda[1] + '\n'
                    #Retroacción correcta
                    if (celda[2]):
                        #Respuesta incorrecta 1. Campo obligatorio
                        pregunta = pregunta + signo_negativo + celda[2] + '\n'
                        #Respuesta incorrecta 2
                        if (celda[3]): pregunta=pregunta + signo_negativo + celda[3] + '\n'
                        #Respuesta incorrecta 3
                        if (celda[4]): pregunta=pregunta + signo_negativo + celda[4] + '\n'
                        #Retroacción general
                        if (celda[5]):
                            pregunta = pregunta + ' ####' + celda[5] + ' \n'

                        #Espacio final
                        pregunta = pregunta + '}\n\n'
                        resultFile.write(pregunta)
                        cuadro_texto.insert(END,u'\n     Pregunta '+str(row-2).zfill(2)+' completada')
                    else: cuadro_texto.insert(END,u'\nLa linea ' + str(row - 2) + u' no contiene una pregunta con la sintaxis correcta')
                else: cuadro_texto.insert(END,u'\nLa linea ' + str(row - 2) + u' no contiene una pregunta con la sintaxis correcta')
            else: cuadro_texto.insert(END,u'\nLa linea ' + str(row-2).zfill(2)+ u' no contiene una pregunta con la sintaxis correcta')

    #Análisis de la segunda pestaña
    #   preguntas de tipo numérico
    try:
        sheet = wb['valor_numerico']
    except KeyError:
        cuadro_texto.insert(END,u'\nNo existe la pestaña \"valor_numerico\", no se añadirán preguntas de este tipo')
    else:

        cuadro_texto.insert(END,u'\n\nLeyendo preguntas de valor numérico')

        for row in range(2, sheet.max_row + 1):
            # En cada fila, vamos leyendo las diferentes celdas
            celda=[]
            #memorizamos toda la fila
            for column in range(1, 7):
                celda.append((sheet.cell(row,column).value))
            if (celda[4]):
                resultFile.write('$CATEGORY: $course$/' + tema + '/' + celda[4] + '\n\n')
            if (celda[0]):
            #Pregunta. Campo obligatoria
                pregunta='::_numerico_' + str(row-1).zfill(2) + '_' + celda[0][0:9] + '::' + celda[0] + ' {#\n'
                #Respuesta correcta. Campo obligatorio
                if isinstance(celda[1], (int, float, complex)):
                    pregunta = pregunta + ' =%100%' + str('{0:.4g}'.format(celda[1]))
                    #Margen de error.
                    if (celda[2]): pregunta = pregunta + ':' + str('{0:.4g}'.format(celda[2]))
                    else: pregunta = pregunta + ':0'
                    #Retroacción correcta
                    if (celda[3]): pregunta = pregunta + '\n ####' + celda[3] + '\n'
                    else: pregunta = pregunta + '\n'
                    #Espacio final
                    pregunta = pregunta + '}\n\n'
                    resultFile.write(pregunta)
                    cuadro_texto.insert(END,'\n     Pregunta '+str(row-1).zfill(2)+' completada')
                else: cuadro_texto.insert(END,u'\nLa linea ' + str(row-1).zfill(2) +u' no contiene una pregunta con la sintaxis correcta')
            else: cuadro_texto.insert(END,u'\nLa linea ' + str(row-1).zfill(2) +u' no contiene una pregunta con la sintaxis correcta')

    #Análisis de la tercera pestaña
    #   preguntas de rellenar huecos
    try:
        sheet = wb['rellenar_huecos']
    except KeyError:
        cuadro_texto.insert(END,u'\nNo existe la pestaña \"rellenar_huecos\", no se añadirán preguntas de este tipo')
    else:

        cuadro_texto.insert(END,u'\n\nLeyendo preguntas de rellenar huecos')

        for row in range(2, sheet.max_row + 1):
            # En cada fila, vamos leyendo las diferentes celdas
            celda=[]
            #memorizamos toda la fila
            for column in range(1, 8):
                celda.append((sheet.cell(row,column).value))
            if (celda[6]):
                resultFile.write('$CATEGORY: $course$/' + tema + '/' + celda[6] + '\n\n')
            if (celda[0]):
                #Si hay retroacción, la memorizamos
                if (celda[5]): retroaccion=' ####' + celda[5] + '\n'
                else: retroaccion=''
            #Pregunta. Campo obligatoria
                pregunta='::_huecos_' + str(row-1).zfill(2)+ '_' + celda[0][0:9] + '::' + celda[0] + ' {\n'
                #Respuesta correcta. Campo obligatorio
                if (celda[1]):
                    pregunta = pregunta + ' =%100%' + str(celda[1]) + '\n'
                    #Respuesta alternativa 1
                    if (celda[2]): pregunta = pregunta + ' =%100%' + str(celda[2]) + '\n'
                    # Respuesta alternativa 2
                    if (celda[3]): pregunta = pregunta + ' =%100%' + str(celda[3]) + '\n'
                    # Respuesta alternativa 3
                    if (celda[4]): pregunta = pregunta + ' =%100%' + str(celda[4]) + '\n'
                    #Espacio final
                    pregunta = pregunta + retroaccion + '}\n\n'
                    resultFile.write(pregunta)
                    cuadro_texto.insert(END,'\n     Pregunta '+str(row-1).zfill(2)+' completada')
                else: cuadro_texto.insert(END,u'\nLa linea ' + str(row-1).zfill(2) +u' no contiene una pregunta con la sintaxis correcta')
            else: cuadro_texto.insert(END,u'\nLa linea ' + str(row-1).zfill(2) +u' no contiene una pregunta con la sintaxis correcta')

    #Análisis de la cuarta pestaña
    #   preguntas de verdadero_falso
    try:
        sheet = wb['verdadero_falso']
    except KeyError:
        cuadro_texto.insert(END,u'\nNo existe la pestaña \"verdadero_falso\", no se añadirán preguntas de este tipo')
    else:

        cuadro_texto.insert(END,u'\n\nLeyendo preguntas de verdadero/falso')

        for row in range(2, sheet.max_row + 1):
            # En cada fila, vamos leyendo las diferentes celdas
            celda=[]
            #memorizamos toda la fila
            for column in range(1, 7):
                celda.append((sheet.cell(row,column).value))
            if (celda[0]):
                #Si hay retroacción, la memorizamos
                if (celda[2]): retroaccion='####' + str(celda[2])
                else: retroaccion=''
                #Pregunta. Campo obligatoria
                pregunta='::_v_f_' + str(row-1).zfill(2)+ '_' + celda[0][0:9] + '::' + celda[0] + ' {'
                #Verdadero o falso. Obligatorio rellenar una u otra
                #if ((celda[1]=='x') and (celda[2]=='')):
                if (celda[1]):
                    if (celda[1]== 'v'):
                        pregunta = pregunta + 'T' + retroaccion + '}\n\n'
                        cuadro_texto.insert(END,'\n     Pregunta ' + str(row-1).zfill(2) + ' completada')
                    elif (celda[1]=='f'):
                        pregunta = pregunta + 'F' + retroaccion + '}\n\n'
                        cuadro_texto.insert(END,'\n     Pregunta ' + str(row-1).zfill(2) + ' completada')
                    else:
                        cuadro_texto.insert(END,u'\nLa linea ' + str(row-1).zfill(2) +u' no contiene una pregunta con la sintaxis correcta')
                        pregunta=''
                else:
                    cuadro_texto.insert(END,u'\nLa linea ' + str(row-1).zfill(2) +u' no contiene una pregunta con la sintaxis correcta')
                    pregunta=''
            else: cuadro_texto.insert(END,u'\nLa linea ' + str(row-1).zfill(2) +u' no contiene una pregunta con la sintaxis correcta')
            resultFile.write(pregunta)

    #Análisis de la quinta pestaña
    #   preguntas de emparejar
    try:
        sheet = wb['emparejar']
    except KeyError:
        cuadro_texto.insert(END,u'\nNo existe la pestaña \"emparejar\", no se añadirán preguntas de este tipo')
    else:

        cuadro_texto.insert(END,u'\n\nLeyendo preguntas de emparejar')

        for row in range(2, sheet.max_row + 1):
            # En cada fila, vamos leyendo las diferentes celdas
            celda=[]
            #memorizamos toda la fila
            for column in range(1, 13):
                celda.append((sheet.cell(row,column).value))
            if (celda[0]):
                #Si hay retroacción, la memorizamos
                if (celda[11]): retroaccion=' ####' + str(celda[11])+'\n'
                else: retroaccion=''
                #Pregunta. Campo obligatoria
                pregunta='::_emparejar_' + str(row-1).zfill(2)+ '_' + celda[0][0:9] + '::' + celda[0] + ' {\n'
                #Primera y segunda opciones con sus respuestas. Obligatorias.
                if (celda[1] and celda[2] and celda[3] and celda[4]):
                    pregunta=pregunta + ' =' + celda[1] + ' -> ' + celda[2] + '\n'
                    pregunta=pregunta + ' =' + celda[3] + ' -> ' + celda[4] + '\n'
                    #Hay una tercera opción?
                    if (celda[5] and celda[6]): pregunta=pregunta + ' =' + celda[5] + ' -> ' + celda[6] + '\n'
                    #Hay una cuarta opción?
                    if (celda[7] and celda[8]): pregunta=pregunta + ' =' + celda[7] + ' -> ' + celda[8] + '\n'
                    # Hay una quinta opción?
                    if (celda[9] and celda[10]): pregunta = pregunta + ' =' + celda[9] + ' -> ' + celda[10] + '\n'
                    pregunta=pregunta+retroaccion+'}\n\n'
                    resultFile.write(pregunta)
                    cuadro_texto.insert(END,'\n     Pregunta ' + str(row-1).zfill(2) + ' completada')

                else: cuadro_texto.insert(END,u'\nLa linea ' + str(row-1).zfill(2) +u' no contiene una pregunta con la sintaxis correcta')
            else: cuadro_texto.insert(END,u'\nLa linea ' + str(row-1).zfill(2) +u' no contiene una pregunta con la sintaxis correcta')

    #Finalización del archivo y del algoritmo
    resultFile.close()
    cuadro_texto.insert(END,'\n\nDone.')

def callback(url):
    webbrowser.open_new(url)

def main():
    #Crear la raíz, donde irán todos los componentes
    raiz=Tk()
    raiz.title("Creador de cuestionarios Moodle")
    raiz.resizable(False,False)
    #raiz.geometry("650x650") No hace falta, se adapta al frame

    #Crear un frame sobre la raíz
    miframe=Frame()
    miframe.pack()
    #miframe.config(width="650", height="650")

    #Crear los diferentes widgets

    #---PRIMERA PARTE: DESCARGAR EL EXCEL VACIO---#
    label_link = Label(miframe, text="PRIMERA PARTE: DESCARGAR LA PLANTILLA:")
    label_link.grid(row=0, column=0, columnspan="2", pady="5")
    #boton descargar
    link1 = Label(miframe, text="Descargar aquí", fg="blue", cursor="hand2")
    link1.grid(row=1,column=0, columnspan="2", pady="5")
    link1.bind("<Button-1>", lambda e: callback("https://github.com/mlvillarroya/moodle-questions/raw/master/excel_moodle.xlsx"))

    #---SEGUNDA PARTE: SUBIR EL ARCHIVO RELLENO---#
    label_link = Label(miframe, text="SEGUNDA PARTE: SUBIR EL ARCHIVO RELLENO:")
    label_link.grid(row=2, column=0, columnspan="2", pady="5")

    #botón para cargar archivo
    boton = Button(miframe, text="Buscar archivo", command=lambda:load_file(ventana))
    boton.grid(row=3,column=0,pady="5")

    def load_file(cuadro):
        fname = askopenfilename(filetypes=(("Excel file", "*.xlsx"),))
        if fname:
            try:
                ruta_archivo.config(text=fname)
                cuadro.insert(END, "Archivo cargado correctamente\n")
            except:  # <- naked except is a bad idea
                showerror("Open Source File", "Failed tofruski read file\n'%s'" % fname)
            return

    #entry text
    ruta_archivo=Label(miframe,width="30", text="archivo.xlsx")
    ruta_archivo.grid(row=3,column=1,sticky="e",padx=10,pady="10")
    ruta_archivo.config(bg="White")

    #---TERCERA PARTE: PREGUNTAS INCORRECTAS RESTAN?---#
    #restan_text = Label(miframe, text="Porcentaje que restan las preguntas incorrectas")
    #restan_text.grid(row=4,column=0,pady="5",padx="10")
    #botón para cargar archivo
    #porcentaje = Entry(miframe)
    #porcentaje.grid(row=4,column=1,pady="5")
    #---CUARTA PARTE: EJECUTAR EL SCRIPT---#
    label_link = Label(miframe, text="CUARTA PARTE: EJECUTAR EL SCRIPT:")
    label_link.grid(row=6, column=0, columnspan="2", pady="5")
    #botón para cargar archivo
    #porciento=porcentaje.get()
    porciento=None
    boton_ejecutar = Button(miframe, text="Ejecutar", command=lambda:generar(ruta_archivo.cget("text"),ventana,porciento))
    boton_ejecutar.grid(row=7,column=0,pady="5", columnspan="2")


    #ventana
    ventana=Text(miframe)
    ventana.grid(row=8,column=0,columnspan=2,padx="10", pady="10")
    ventana.config(width="50")

    #comenzar la interfaz
    raiz.mainloop()

main()